from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.core.validators import EmailValidator, ValidationError as DjangoValidationError
from django.db import transaction
from openpyxl import load_workbook
from staffs.models import Staff
import re


HEADER_MAPS = [
    {"first_name": "first_name", "last_name": "last_name", "email": "email"},
    {"first_name": "نام", "last_name": "نام خانوادگی", "email": "ایمیل"},
]

_email_validator = EmailValidator()

AT_PATTERNS = [
    r"\s*\[\s*at\s*\]\s*", r"\s*\(\s*at\s*\)\s*", r"\s*\{\s*at\s*\}\s*",
    r"\s+at\s+", r"\sat\s", r" at ",
]
DOT_PATTERNS = [
    r"\s*\[\s*dot\s*\]\s*", r"\s*\(\s*dot\s*\)\s*", r"\s*\{\s*dot\s*\}\s*",
    r"\s+dot\s+", r"\sdot\s", r" dot ",
    r"\s*\[\s*\.\s*\]\s*", r"\s*\(\s*\.\s*\)\s*", r"\s*\{\s*\.\s*\}\s*",
]

SPLIT_REGEX = re.compile(r"[,\;/\s]+")

CLEAN_CHARS_REGEX = re.compile(r"[^a-zA-Z0-9._%+\-@]+")

def normalize_raw_email(raw: str) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()

    if not s:
        return None

    s = s.replace("＠", "@")

    low = s.lower()

    for pat in AT_PATTERNS:
        low = re.sub(pat, "@", low, flags=re.IGNORECASE)

    for pat in DOT_PATTERNS:
        low = re.sub(pat, ".", low, flags=re.IGNORECASE)

    low = low.replace("[at]", "@").replace("(at)", "@").replace("{at}", "@")
    low = low.replace("[dot]", ".").replace("(dot)", ".").replace("{dot}", ".")

    low = re.sub(r"\s*@\s*", "@", low)
    low = re.sub(r"\s*\.\s*", ".", low)

    low = CLEAN_CHARS_REGEX.sub("", low)

    if low.count("@") > 1:
        first = low.split("@", 1)
        right = first[1].replace("@", "")
        low = first[0] + "@" + right

    return low or None


def pick_best_email(candidate_field: str) -> str | None:
    
    if not candidate_field:
        return None

    parts = [p for p in SPLIT_REGEX.split(candidate_field) if p]
    normalized_valids = []

    for part in parts:
        norm = normalize_raw_email(part)
        if not norm:
            continue
        try:
            _email_validator(norm)
            normalized_valids.append(norm)
        except DjangoValidationError:
            continue

    if not normalized_valids:
        return None

    def score(email: str) -> tuple:
        domain = email.split("@")[-1].lower()
        if domain.endswith("ui.ac.ir"):
            return (0, email)
        if domain.endswith(".ac.ir"):
            return (1, email)
        return (2, email)

    normalized_valids.sort(key=score)
    return normalized_valids[0]


def normalize_username_from_email(email: str) -> str | None:
    if not email or "@" not in email:
        return None
    base = email.split("@")[0].lower()
    base = re.sub(r"[^a-z0-9._-]+", "", base)
    return base or None


class Command(BaseCommand):
    help = "Import staffs from an Excel file (.xlsx). Example: manage.py import_staffs /path/to/file.xlsx [--update] [--sheet Sheet1]"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to .xlsx file (container path if using Docker)")
        parser.add_argument(
            "--update",
            action="store_true",
            help="If given, update existing Staff/User by email (case-insensitive) instead of skipping."
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Optional sheet name. If omitted, the first sheet will be used."
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = options["xlsx_path"]
        update_existing = options["update"]
        sheet_name = options["sheet"]

        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Cannot open Excel file: {e}")

        try:
            ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        except Exception as e:
            raise CommandError(f"Cannot open worksheet: {e}")

        headers = []
        for cell in ws[1]:
            v = cell.value
            headers.append((v.strip() if isinstance(v, str) else v))

        header_map = None
        for m in HEADER_MAPS:
            if all(mv in headers for mv in m.values()):
                header_map = m
                break

        if not header_map:
            raise CommandError(
                f"Cannot detect headers. Got: {headers}\n"
                f"Expected one of maps like: {HEADER_MAPS}"
            )

        idx = {key: headers.index(colname) for key, colname in header_map.items()}

        created, updated, skipped, errors = 0, 0, 0, 0
        row_no = 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_no += 1
            try:
                raw_first_name = row[idx["first_name"]]
                raw_last_name  = row[idx["last_name"]]
                raw_email      = row[idx["email"]]

                first_name = (raw_first_name or "").strip() if isinstance(raw_first_name, str) else (str(raw_first_name).strip() if raw_first_name else "")
                last_name  = (raw_last_name or "").strip() if isinstance(raw_last_name, str) else (str(raw_last_name).strip() if raw_last_name else "")
                email_cell = (raw_email or "").strip()

                best_email = pick_best_email(email_cell)
                if not best_email:
                    skipped += 1
                    self.stdout.write(self.style.WARNING(f"Row {row_no}: invalid email '{email_cell}' -> skipped"))
                    continue

                staff = Staff.objects.filter(email__iexact=best_email).first()

                if staff:
                    if update_existing:
                        if first_name:
                            staff.first_name = first_name
                        if last_name:
                            staff.last_name = last_name
                        staff.email = best_email
                        staff.save()

                        if staff.user_id:
                            u = staff.user
                            u.first_name = staff.first_name
                            u.last_name = staff.last_name
                            u.email = staff.email
                            u.save()

                        updated += 1
                    else:
                        skipped += 1
                    continue

                username = normalize_username_from_email(best_email) or f"user_{best_email.split('@')[0]}"

                base_username = username
                suf = 1
                while User.objects.filter(username=username).exists():
                    suf += 1
                    username = f"{base_username}{suf}"

                user = User(
                    username=username,
                    email=best_email,
                    first_name=first_name,
                    last_name=last_name,
                )
                user.set_unusable_password()
                user.save()

                Staff.objects.create(
                    user=user,
                    first_name=first_name,
                    last_name=last_name,
                    email=best_email,
                    is_registered=False,
                )
                created += 1

            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_no}: error -> {e}"))

        self.stdout.write(self.style.SUCCESS(
            f"Done. created={created}, updated={updated}, skipped={skipped}, errors={errors}"
        ))
